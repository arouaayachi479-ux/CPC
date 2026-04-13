import streamlit as st
import pandas as pd
import io
import unicodedata
import re
from openpyxl import load_workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

# ─── Page config ────────────────────────────────────────────────────────────────
st.set_page_config(page_title="CPC Builder · LabelVie", page_icon="🏪", layout="wide")

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&family=Poppins:wght@600;700&display=swap');

html, body, [class*="css"] {
    font-family: 'Inter', sans-serif;
}

/* ── Background ── */
.stApp {
    background: #f5f6fa;
    color: #1a1a2e;
}

/* ── Typography ── */
h1, h2, h3 {
    font-family: 'Poppins', sans-serif !important;
    color: #1a1a2e !important;
}

/* ── Top banner stripe ── */
.top-banner {
    background: linear-gradient(90deg, #E30613 0%, #b30000 60%, #007a3d 100%);
    border-radius: 14px;
    padding: 1.4rem 2rem;
    margin-bottom: 1.8rem;
    display: flex;
    align-items: center;
    gap: 1rem;
}
.top-banner h1 {
    color: #ffffff !important;
    font-family: 'Poppins', sans-serif !important;
    font-size: 1.7rem !important;
    margin: 0 !important;
    line-height: 1.2;
}
.top-banner p {
    color: rgba(255,255,255,0.82);
    font-size: 0.85rem;
    margin: 4px 0 0 0;
}

/* ── Section title ── */
.block-title {
    font-family: 'Poppins', sans-serif;
    font-size: 0.72rem;
    letter-spacing: 0.14em;
    text-transform: uppercase;
    color: #E30613;
    border-left: 3px solid #E30613;
    padding-left: 10px;
    margin-bottom: 0.9rem;
    font-weight: 600;
}
.block-title.green {
    color: #00A651;
    border-left-color: #00A651;
}

/* ── Cards ── */
.upload-card {
    background: #ffffff;
    border: 1px solid #e0e4ed;
    border-top: 3px solid #E30613;
    border-radius: 12px;
    padding: 1.4rem;
    margin-bottom: 1rem;
    box-shadow: 0 2px 8px rgba(0,0,0,0.06);
}
.upload-card.green {
    border-top-color: #00A651;
}

/* ── Info pills ── */
.info-pill {
    display: inline-block;
    background: #f0f2f8;
    border: 1px solid #dde1ec;
    border-radius: 20px;
    padding: 3px 12px;
    font-size: 0.76rem;
    color: #4a4f6a;
    margin: 3px;
    font-weight: 500;
}

/* ── Metric boxes ── */
.metric-box {
    background: #ffffff;
    border: 1px solid #e0e4ed;
    border-radius: 12px;
    padding: 1.1rem 1.5rem;
    text-align: center;
    box-shadow: 0 1px 6px rgba(0,0,0,0.05);
}
.metric-num {
    font-family: 'Poppins', sans-serif;
    font-size: 2.1rem;
    color: #1a1a2e;
    line-height: 1;
    font-weight: 700;
}
.metric-num.red   { color: #E30613; }
.metric-num.green { color: #00A651; }
.metric-label {
    font-size: 0.73rem;
    color: #8a8fa8;
    text-transform: uppercase;
    letter-spacing: 0.1em;
    margin-top: 5px;
}

/* ── Buttons ── */
.stButton > button {
    background: linear-gradient(135deg, #E30613, #b30000) !important;
    color: #ffffff !important;
    font-family: 'Poppins', sans-serif !important;
    font-weight: 600 !important;
    border: none !important;
    border-radius: 8px !important;
    padding: 0.65rem 2.2rem !important;
    letter-spacing: 0.03em !important;
    box-shadow: 0 3px 10px rgba(227,6,19,0.28) !important;
    transition: all .2s !important;
}
.stButton > button:hover {
    background: linear-gradient(135deg, #ff1a27, #cc0000) !important;
    box-shadow: 0 5px 14px rgba(227,6,19,0.38) !important;
}

/* ── Download button ── */
[data-testid="stDownloadButton"] > button {
    background: linear-gradient(135deg, #00A651, #007a3d) !important;
    color: #ffffff !important;
    font-family: 'Poppins', sans-serif !important;
    font-weight: 600 !important;
    border: none !important;
    border-radius: 8px !important;
    padding: 0.65rem 2.2rem !important;
    box-shadow: 0 3px 10px rgba(0,166,81,0.28) !important;
}
[data-testid="stDownloadButton"] > button:hover {
    background: linear-gradient(135deg, #00c060, #009040) !important;
}

/* ── File uploader ── */
[data-testid="stFileUploadDropzone"] {
    background: #fafbfd !important;
    border: 1.5px dashed #c8cdd9 !important;
    border-radius: 8px !important;
}

/* ── Banners ── */
.success-banner {
    background: linear-gradient(135deg, #e8f8ef, #d0f0e0);
    border: 1px solid #00A651;
    border-left: 4px solid #00A651;
    border-radius: 10px;
    padding: 1rem 1.5rem;
    color: #006b35;
    font-weight: 600;
    font-size: 0.9rem;
    margin: 1rem 0;
}
.warn-banner {
    background: linear-gradient(135deg, #fff8e6, #fff0cc);
    border: 1px solid #e6a817;
    border-left: 4px solid #e6a817;
    border-radius: 10px;
    padding: 0.9rem 1.4rem;
    color: #7a5a00;
    font-size: 0.84rem;
    margin: 0.5rem 0;
}
.empty-state {
    text-align: center;
    padding: 3.5rem 2rem;
    color: #a0a8bc;
    font-size: 0.88rem;
    background: #ffffff;
    border-radius: 12px;
    border: 1px dashed #d0d4e0;
}

/* ── Separator ── */
.sep {
    border: none;
    border-top: 1px solid #e4e7f0;
    margin: 2rem 0;
}

/* ── Dataframe ── */
[data-testid="stDataFrame"] {
    border-radius: 10px !important;
    overflow: hidden;
    border: 1px solid #e0e4ed !important;
}
</style>
""", unsafe_allow_html=True)

# ─── Helpers ────────────────────────────────────────────────────────────────────

def normalize(text):
    if text is None:
        return ""
    text = str(text).strip()
    text = unicodedata.normalize("NFD", text)
    text = "".join(c for c in text if unicodedata.category(c) != "Mn")
    return re.sub(r"\s+", " ", text).lower().strip()


@st.cache_data(show_spinner=False)
def load_check_alcool(raw_bytes):
    xl = pd.ExcelFile(io.BytesIO(raw_bytes))
    sheet = next(
        (s for s in xl.sheet_names if normalize(s) == "check alcool"),
        None,
    )
    if sheet is None:
        return None, f"Feuille 'check alcool' introuvable. Feuilles dispo : {xl.sheet_names}"
    df = xl.parse(sheet, header=0)
    df.columns = [str(c).strip() for c in df.columns]
    return df, None


@st.cache_data(show_spinner=False)
def load_cpc_sheets_list(raw_bytes):
    wb = load_workbook(io.BytesIO(raw_bytes), read_only=True, data_only=True)
    names = wb.sheetnames
    wb.close()
    return names


def extract_one_sheet(raw_bytes, sheet_name):
    wb = load_workbook(io.BytesIO(raw_bytes), read_only=True, data_only=True)
    ws = wb[sheet_name]
    code_site = None
    rows_data = []
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if i == 2 and len(row) > 1:
            code_site = str(row[1]).strip() if row[1] is not None else None
        if i < 8:
            continue
        if len(row) < 31:
            continue
        libelle = row[1]
        valeur  = row[30]
        if libelle is not None and str(libelle).strip():
            rows_data.append((str(libelle).strip(), valeur))
    wb.close()
    return code_site, rows_data


def build_styled_excel(df_result: pd.DataFrame) -> bytes:
    from openpyxl import Workbook

    RED   = "E30613"
    GREEN = "00A651"
    WHITE = "FFFFFF"
    LIGHT_GRAY = "F5F6FA"
    DARK_TEXT  = "1A1A2E"

    thin_side  = Side(style="thin", color="D0D4E0")
    thin_border = Border(left=thin_side, right=thin_side,
                         top=thin_side,  bottom=thin_side)

    wb = Workbook()
    ws = wb.active
    ws.title = "Base CPC"

    ws.freeze_panes = "A3"

    n_cols = len(df_result.columns)
    ws.merge_cells(start_row=1, start_column=1,
                   end_row=1,   end_column=n_cols)
    title_cell = ws.cell(row=1, column=1,
                         value="BASE CPC CONSOLIDÉE — LabelVie")
    title_cell.font      = Font(name="Arial", bold=True, size=13, color=WHITE)
    title_cell.fill      = PatternFill("solid", fgColor=GREEN)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    title_cell.border    = thin_border
    ws.row_dimensions[1].height = 28

    for col_idx, col_name in enumerate(df_result.columns, start=1):
        cell = ws.cell(row=2, column=col_idx, value=col_name)
        cell.font      = Font(name="Arial", bold=True, size=10, color=WHITE)
        cell.fill      = PatternFill("solid", fgColor=RED)
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
        cell.border    = thin_border
    ws.row_dimensions[2].height = 22

    for row_idx, row in enumerate(df_result.itertuples(index=False), start=3):
        is_even = (row_idx % 2 == 0)
        row_fill = PatternFill("solid", fgColor=LIGHT_GRAY) if is_even else None

        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font      = Font(name="Arial", size=9, color=DARK_TEXT)
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border    = thin_border
            if row_fill:
                cell.fill = row_fill

        ws.row_dimensions[row_idx].height = 16

    cumule_idx = None
    for i, col in enumerate(df_result.columns, start=1):
        if "cumul" in normalize(col):
            cumule_idx = i
            break
    if cumule_idx:
        for row_idx in range(3, len(df_result) + 3):
            cell = ws.cell(row=row_idx, column=cumule_idx)
            cell.number_format = "#,##0.00"
            cell.alignment     = Alignment(horizontal="right", vertical="center")

    total_row = len(df_result) + 3
    for col_idx in range(1, n_cols + 1):
        col_name = df_result.columns[col_idx - 1]
        cell = ws.cell(row=total_row, column=col_idx)
        if col_idx == 1:
            cell.value = "TOTAL"
            cell.font  = Font(name="Arial", bold=True, size=9, color=WHITE)
        elif "cumul" in normalize(col_name):
            cell.value = f"=SUM({get_column_letter(col_idx)}3:{get_column_letter(col_idx)}{total_row-1})"
            cell.font  = Font(name="Arial", bold=True, size=9, color=WHITE)
            cell.number_format = "#,##0.00"
            cell.alignment = Alignment(horizontal="right", vertical="center")
        else:
            cell.font = Font(name="Arial", bold=True, size=9, color=WHITE)
        cell.fill   = PatternFill("solid", fgColor=GREEN)
        cell.border = thin_border
        ws.row_dimensions[total_row].height = 18

    col_widths = {
        "Code Site":   13,
        "Lib Magasin": 22,
        "Poste":       40,
        "Cumulé":      16,
        "Surface":     12,
        "Alcool":      10,
    }
    for col_idx, col_name in enumerate(df_result.columns, start=1):
        width = col_widths.get(col_name, 18)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.auto_filter.ref = (
        f"A2:{get_column_letter(n_cols)}{len(df_result) + 2}"
    )

    ws.oddFooter.center.text = (
        "&\"Arial,Italic\"&8LabelVie — CPC Builder | "
        "Page &P / &N"
    )
    ws.oddFooter.center.color = "888888"

    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage   = True
    ws.page_setup.fitToWidth  = 1
    ws.sheet_view.showGridLines = False

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ─── Colonnes fixes ──────────────────────────────────────────────────────────────
COL_CODE    = "Code"
COL_MAG     = "Mag"
COL_SURFACE = "Surface"
COL_ALCOOL  = "Alcool"

# ─── Header ─────────────────────────────────────────────────────────────────────
st.markdown("""
<div class="top-banner">
  <div>
    <h1>🏪 CPC Builder</h1>
    <p>Consolidation automatique des données CPC par magasin — LabelVie</p>
  </div>
</div>
""", unsafe_allow_html=True)

# ─── Upload ──────────────────────────────────────────────────────────────────────
col1, col2 = st.columns(2, gap="large")

with col1:
    st.markdown('<div class="block-title">① Check Alcool</div>', unsafe_allow_html=True)
    st.markdown('<div class="upload-card">', unsafe_allow_html=True)
    file_check = st.file_uploader(
        "Fichier contenant la feuille **check alcool**",
        type=["xlsx", "xls"], key="check"
    )
    st.markdown(
        '<span class="info-pill">Code</span>'
        '<span class="info-pill">Mag</span>'
        '<span class="info-pill">Surface</span>'
        '<span class="info-pill">Alcool</span>',
        unsafe_allow_html=True)
    st.markdown('</div>', unsafe_allow_html=True)

with col2:
    st.markdown('<div class="block-title green">② CPC Total</div>', unsafe_allow_html=True)
    st.markdown('<div class="upload-card green">', unsafe_allow_html=True)
    file_cpc = st.file_uploader(
        "Fichier CPC Total (1 feuille = 1 magasin)",
        type=["xlsx", "xls"], key="cpc"
    )
    st.markdown(
        '<span class="info-pill">Code site → B2</span>'
        '<span class="info-pill">Poste → col B</span>'
        '<span class="info-pill">Cumulé → col AE</span>',
        unsafe_allow_html=True)
    st.markdown('</div>', unsafe_allow_html=True)

st.markdown("<hr class='sep'>", unsafe_allow_html=True)

# ─── Guard ──────────────────────────────────────────────────────────────────────
if not (file_check and file_cpc):
    st.markdown(
        "<div class='empty-state'>"
        "⬆️&nbsp; Chargez les deux fichiers pour commencer la consolidation"
        "</div>",
        unsafe_allow_html=True)
    st.stop()

bytes_check = file_check.read()
bytes_cpc   = file_cpc.read()

# ── 1. Check alcool ─────────────────────────────────────────────────────────────
with st.spinner("Lecture du fichier check alcool…"):
    df_check, err = load_check_alcool(bytes_check)

if err:
    st.error(f"❌ {err}")
    st.stop()

if COL_CODE not in df_check.columns:
    st.error(
        f"❌ Colonne obligatoire **'{COL_CODE}'** introuvable.\n\n"
        f"Colonnes détectées : `{list(df_check.columns)}`"
    )
    st.stop()

# ── Colonnes optionnelles ────────────────────────────────────────────────────
has_mag     = COL_MAG     in df_check.columns
has_surface = COL_SURFACE in df_check.columns
has_alcool  = COL_ALCOOL  in df_check.columns

# Feedback sur les colonnes détectées
pills_html = '<span class="info-pill" style="background:#e8f8ef;border-color:#00A651;color:#006b35">✅ Code</span>'
for col, flag, label in [
    (COL_MAG,     has_mag,     "Mag"),
    (COL_SURFACE, has_surface, "Surface"),
    (COL_ALCOOL,  has_alcool,  "Alcool"),
]:
    if flag:
        pills_html += f'<span class="info-pill" style="background:#e8f8ef;border-color:#00A651;color:#006b35">✅ {label}</span>'
    else:
        pills_html += f'<span class="info-pill" style="background:#fff8e6;border-color:#e6a817;color:#7a5a00">⚠️ {label} (absent)</span>'

st.markdown(pills_html, unsafe_allow_html=True)

if not has_mag:
    st.markdown(
        "<div class='warn-banner'>ℹ️ Colonne <b>Mag</b> absente — "
        "le matching se fera par <b>Code</b> (comparé au code site B2 de chaque feuille CPC).</div>",
        unsafe_allow_html=True)

# ── Aperçu colonnes disponibles ───────────────────────────────────────────────
preview_cols = [c for c in [COL_CODE, COL_MAG, COL_SURFACE, COL_ALCOOL]
                if c in df_check.columns]
st.markdown('<div class="block-title">Aperçu — Check Alcool</div>',
            unsafe_allow_html=True)
st.dataframe(df_check[preview_cols].head(10), use_container_width=True)

@st.cache_data(show_spinner=False)
def get_all_sheet_codes(raw_bytes, sheet_names):
    """Lit uniquement la cellule B2 de chaque feuille pour récupérer le code site."""
    result = {}
    wb = load_workbook(io.BytesIO(raw_bytes), read_only=True, data_only=True)
    for name in sheet_names:
        try:
            ws = wb[name]
            code = None
            for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
                if i == 2:
                    code = str(row[1]).strip() if len(row) > 1 and row[1] is not None else None
                    break
            result[name] = code
        except Exception:
            result[name] = None
    wb.close()
    return result  # {sheet_name: code_site}
# ── 2. Feuilles CPC ─────────────────────────────────────────────────────────────
with st.spinner("Lecture des feuilles CPC…"):
    sheets_cpc = load_cpc_sheets_list(bytes_cpc)

# ── 3. Matching ─────────────────────────────────────────────────────────────────
matched, unmatched = [], []

if has_mag:
    # ── Matching par nom de feuille ↔ colonne Mag ──────────────────────────────
    sheet_norm = {normalize(s): s for s in sheets_cpc}
    magasins   = df_check[COL_MAG].dropna().astype(str).str.strip().unique().tolist()

    for mag in magasins:
        found = sheet_norm.get(normalize(mag))
        if found:
            matched.append((mag, found))   # (mag_name, sheet_name)
        else:
            unmatched.append(mag)

    match_mode = "mag"
    identifiers = magasins

else:
    # ── Matching par Code ↔ code site B2 de chaque feuille ────────────────────
    with st.spinner("Lecture des codes site dans les feuilles CPC…"):
        sheet_codes = get_all_sheet_codes(bytes_cpc, tuple(sheets_cpc))

    # Inverser : code_site → sheet_name (premier trouvé si doublons)
    code_to_sheet = {}
    for sheet_name, code in sheet_codes.items():
        if code and code not in code_to_sheet:
            code_to_sheet[code] = sheet_name

    codes_check = df_check[COL_CODE].dropna().astype(str).str.strip().unique().tolist()

    for code in codes_check:
        found = code_to_sheet.get(code)
        if found:
            matched.append((code, found))   # (code, sheet_name)
        else:
            unmatched.append(code)

    match_mode  = "code"
    identifiers = codes_check

# ── Métriques ────────────────────────────────────────────────────────────────────
mc1, mc2, mc3 = st.columns(3)
with mc1:
    label_total = "Magasins Check" if has_mag else "Codes Check"
    st.markdown(
        f'<div class="metric-box">'
        f'<div class="metric-num">{len(identifiers)}</div>'
        f'<div class="metric-label">{label_total}</div></div>',
        unsafe_allow_html=True)
with mc2:
    st.markdown(
        f'<div class="metric-box">'
        f'<div class="metric-num green">{len(matched)}</div>'
        f'<div class="metric-label">Feuilles trouvées</div></div>',
        unsafe_allow_html=True)
with mc3:
    st.markdown(
        f'<div class="metric-box">'
        f'<div class="metric-num red">{len(unmatched)}</div>'
        f'<div class="metric-label">Non trouvés</div></div>',
        unsafe_allow_html=True)

if unmatched:
    label_warn = "Magasins" if has_mag else "Codes"
    st.markdown(
        f"<div class='warn-banner'>⚠️ {label_warn} sans feuille CPC : "
        f"{', '.join(unmatched)}</div>",
        unsafe_allow_html=True)
    if has_mag:
        with st.expander("🔬 Diagnostic repr() — caractères cachés"):
            st.markdown("**Colonne Mag (check alcool) :**")
            for v in identifiers:
                st.code(repr(v))
            st.markdown("**Feuilles CPC :**")
            for s in sheets_cpc:
                st.code(repr(s))
    else:
        with st.expander("🔬 Diagnostic — codes non trouvés"):
            st.markdown("**Codes dans check alcool :**")
            for v in unmatched:
                st.code(repr(v))
            st.markdown("**Codes lus en B2 des feuilles CPC :**")
            for sheet, code in sheet_codes.items():
                st.code(f"{sheet!r}  →  {code!r}")

st.markdown("<hr class='sep'>", unsafe_allow_html=True)

if not matched:
    st.warning("Aucun magasin trouvé. Consultez le diagnostic ci-dessus.")
    st.stop()

# ── 4. Consolidation ────────────────────────────────────────────────────────────
if st.button("🔨  Construire la base consolidée"):

    all_rows = []
    progress = st.progress(0, text="Extraction en cours…")

    for idx, (identifier, sheet_name) in enumerate(matched):
        code_site, postes = extract_one_sheet(bytes_cpc, sheet_name)

        # Trouver la ligne correspondante dans check alcool
        if has_mag:
            mask = df_check[COL_MAG].astype(str).str.strip().str.lower() == identifier.lower()
        else:
            mask = df_check[COL_CODE].astype(str).str.strip() == identifier

        row_check = df_check[mask].iloc[0] if mask.any() else pd.Series(dtype=object)

        code_mag = row_check.get(COL_CODE) if not row_check.empty else code_site
        mag_name = (row_check.get(COL_MAG)     if not row_check.empty else identifier) if has_mag else identifier
        surface  = (row_check.get(COL_SURFACE)  if not row_check.empty else None)      if has_surface else None
        alcool   = (row_check.get(COL_ALCOOL)   if not row_check.empty else None)      if has_alcool  else None

        for poste, cumule in postes:
            row = {
                "Code Site":   code_mag,
                "Lib Magasin": mag_name,
                "Poste":       poste,
                "Cumulé":      cumule,
            }
            if has_surface:
                row["Surface"] = surface
            if has_alcool:
                row["Alcool"] = alcool
            all_rows.append(row)

        progress.progress((idx + 1) / len(matched),
                          text=f"Traitement : {mag_name}")

    progress.empty()

    if not all_rows:
        st.warning("Aucune donnée extraite.")
        st.stop()

    result_cols = ["Code Site", "Lib Magasin", "Poste", "Cumulé"]
    if has_surface:
        result_cols.append("Surface")
    if has_alcool:
        result_cols.append("Alcool")

    df_result = pd.DataFrame(all_rows, columns=result_cols)

    st.markdown(
        "<div class='success-banner'>"
        "✅ Base construite avec succès !</div>",
        unsafe_allow_html=True)
    st.caption(f"{len(df_result):,} lignes × {len(df_result.columns)} colonnes")
    st.dataframe(df_result, use_container_width=True)

    excel_bytes = build_styled_excel(df_result)

    st.download_button(
        label="⬇️  Télécharger la base Excel (LabelVie)",
        data=excel_bytes,
        file_name="base_cpc_consolidee.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
